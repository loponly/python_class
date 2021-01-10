from openpyxl import Workbook, load_workbook

filename = 'Excel_manipulation/book.xlsx'
wb = load_workbook(filename)

print(wb.sheetnames)

wb.active = 0
number = 5
name = 'John'

print('This is a %d \n %s' % (number, name))
print('This is a'+str(number)+'\n '+name+'')
print(f'This is a {number} \n {name}')


ws = wb.active

for cell in ws['A1:ZZ999']:
    if cell[0].value is None:
        break
    print(cell[0].value, cell[1].value, cell[2].value)
    if cell[0].value == 'pants':
        cell[0].value = 'hat'

wb.save(filename)
wb.close()
